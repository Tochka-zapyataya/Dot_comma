"""
Excel «график» смен по дням и станциям: сотрудник, начало/конец/длина + сетка часов с заливкой.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config

logger = logging.getLogger(__name__)

_WEEKDAY_RU = (
    "Понедельник",
    "Вторник",
    "Среда",
    "Четвер",
    "Пятница",
    "Суббота",
    "Воскресенье",
)

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_BLACK = PatternFill(fill_type="solid", start_color="FF000000", end_color="FF000000")
_FILL_BLUE = PatternFill(fill_type="solid", start_color="FF4F81BD", end_color="FF4F81BD")
_FILL_WHITE = PatternFill(fill_type="solid", start_color="FFFFFFFF", end_color="FFFFFFFF")

_FONT_WHITE_BOLD = Font(bold=True, color="FFFFFFFF", size=11)
_FONT_HEADER = Font(bold=True, size=10)
_FONT_NORMAL = Font(size=10)


def _ru_date(ds: str) -> str:
    try:
        d = datetime.strptime(ds, "%Y-%m-%d").date()
    except ValueError:
        return ds
    return d.strftime("%d.%m.%Y")


def _weekday_ru(ds: str) -> str:
    try:
        d = datetime.strptime(ds, "%Y-%m-%d").date()
    except ValueError:
        return ""
    return _WEEKDAY_RU[d.weekday()]


def _hour_labels() -> list[int]:
    return list(config.HOURS)


def _fmt_clock(h: int) -> str:
    return f"{int(h)}:00"


def _fmt_duration_hours(start: int, end: int) -> str:
    d = int(end) - int(start)
    return f"{d:.2f}".replace(".", ",")


def write_schedule_table_xlsx(schedule_df: pd.DataFrame, out_path: Path) -> None:
    """Пишет table.xlsx: блоки «день × станция», строка = одна смена."""
    out_path = Path(out_path)
    hours = _hour_labels()
    if not hours:
        hours = list(range(config.OPEN_HOUR, config.CLOSE_HOUR))
    n_timecols = len(hours)
    total_cols = 4 + n_timecols

    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    row = 1
    if schedule_df is None or schedule_df.empty:
        ws.cell(row=row, column=1, value="Нет данных расписания")
        wb.save(out_path)
        logger.warning("schedule empty — wrote stub %s", out_path)
        return

    df = schedule_df.copy()
    df["ds"] = df["ds"].astype(str)
    df["station_key"] = df["station_key"].astype(str)
    df["employee_id"] = df["employee_id"].astype(int)
    df["starttime"] = df["starttime"].astype(int)
    df["finishtime"] = df["finishtime"].astype(int)

    dates = sorted(df["ds"].unique())
    stations = list(config.STATIONS)

    for ds in dates:
        sub_d = df[df["ds"] == ds]
        for st_key in stations:
            sub = sub_d[sub_d["station_key"] == st_key]
            if sub.empty:
                continue

            day_label = f"{_weekday_ru(ds)} · {_ru_date(ds)}"
            st_title = (config.STATION_NAMES.get(st_key, st_key) or st_key).upper()

            ws.cell(row=row, column=1, value=day_label)
            ws.cell(row=row, column=1).font = Font(size=10, italic=True, color="FF555555")
            row += 1

            top_merge = f"A{row}:{get_column_letter(total_cols)}{row}"
            ws.merge_cells(top_merge)
            c = ws.cell(row=row, column=1, value=st_title)
            c.fill = _FILL_BLACK
            c.font = _FONT_WHITE_BOLD
            c.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(1, total_cols + 1):
                ws.cell(row=row, column=col).border = _BORDER
            row += 1

            headers = ["Сотрудник", "Начало", "Окончание", "Длина", *[str(h) for h in hours]]
            for col, text in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=text)
                cell.font = _FONT_HEADER
                cell.fill = _FILL_WHITE
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER
            row += 1

            block = sub.sort_values(["starttime", "employee_id"])
            for _, r in block.iterrows():
                eid = int(r["employee_id"])
                st_t = int(r["starttime"])
                fi_t = int(r["finishtime"])
                name = f"Сотрудник {eid}"
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=_fmt_clock(st_t))
                ws.cell(row=row, column=3, value=_fmt_clock(fi_t))
                ws.cell(row=row, column=4, value=_fmt_duration_hours(st_t, fi_t))
                for c in range(1, 5):
                    cell = ws.cell(row=row, column=c)
                    cell.font = _FONT_NORMAL
                    cell.border = _BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for idx, h in enumerate(hours):
                    col = 5 + idx
                    cell = ws.cell(row=row, column=col)
                    cell.border = _BORDER
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if st_t <= h < fi_t:
                        cell.fill = _FILL_BLUE
                        cell.value = 1
                        cell.font = Font(size=9, color="FFFFFFFF", bold=True)
                    else:
                        cell.fill = _FILL_WHITE
                        cell.value = ""
                        cell.font = Font(size=9, color="FF333333")

                row += 1

            row += 1

    ws.column_dimensions["A"].width = 16
    for c in range(2, 5):
        ws.column_dimensions[get_column_letter(c)].width = 11
    for idx in range(n_timecols):
        ws.column_dimensions[get_column_letter(5 + idx)].width = 3.2

    wb.save(out_path)
    logger.info("Wrote schedule gantt table %s", out_path)
